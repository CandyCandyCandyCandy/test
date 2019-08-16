from openpyxl import load_workbook
from openpyxl import Workbook
import os
import openpyxl
import os,sys,shutil
import xlwings as xw


Special_title = ['Pass','Fail']




def get_report_content(work_sheet):
    #获取报告内容
    report_content = []
    for r in range(1,work_sheet.max_row+1):
        for c in range(1,work_sheet.max_column+1):
            report_content.append(work_sheet.cell(row = r,column = c).value)
    return report_content



def self_update_ask():
    answer = input('想要同步哪些表格？1：提测产品基本信息，2：里程碑软件测试报告，3：软件测试用例，4：返回\n')
    if answer == '1':
        search = '提测产品基本信息'
        report_name = '03-V2.2 小米生态链提测产品基本信息'
    elif answer == '2':
        search = '里程碑软件测试报告'
        report_name = '05-V1.0小米生态链里程碑软件测试报告'
    elif answer =='3':
        search = '软件测试用例'
        report_name = '04-V1.0小米生态链软件测试用例'
    else:
        print('请选择一个有效输入')
        return
    if answer == '4':
        return
    filenames = os.listdir()
    report_files = []
    for i in filenames:
        if 'xlsx' in i and search in i and '模板' not in i:
            report_files.append(i)
    else:
        total_information_self_update(report_files,search,report_name)


def report_update_ask():
    answer = input('想要同步哪些表格？1：提测产品基本信息，2：里程碑软件测试报告，3：软件测试用例,4：返回\n')
    if answer == '1':
        search = '提测产品基本信息'
        report_name = '03-V2.2 小米生态链提测产品基本信息'
    elif answer == '2':
        search = '里程碑软件测试报告'
        report_name = '05-V1.0小米生态链里程碑软件测试报告'
    elif answer =='3':
        search = '软件测试用例'
        report_name = '04-V1.0小米生态链软件测试用例'
    else:
        print('请选择一个有效输入')
        return
    if answer == '4':
        return
    filenames = os.listdir()
    report_files = []
    for i in filenames:
        if 'xlsx' in i and search in i and '模板' not in i:
            report_files.append(i)
    if answer == '2':
        report_2_update(report_files,search,report_name)
    else:
        report_update(report_files,search,report_name)
        

def total_information_self_update(report_files,name,report_name):
    #更新至Totol的信息表中
    total_report = load_workbook('Total_Information.xlsx')
    sheet_names = total_report.get_sheet_names()
    total_sheet = total_report.get_sheet_by_name(name+'_所有报告信息')
    titles = []
    for i in range(1,total_sheet.max_column+1):
        title,judge = extract_content(total_sheet.cell(row=1, column = i).value)
        titles.append(title)
    dummy_report = load_workbook('IoT-QA-D-'+report_name+'模板dummy.xlsx')
    dummy_sheets = dummy_report.get_sheet_names()[1]
    dummy_sheet = dummy_report.get_sheet_by_name(dummy_sheets)
    print(titles)
    coordinate = get_content_coordinate(titles[1:],dummy_sheet)
    row_index = 2
    for i in report_files:
        total_sheet.cell(row = row_index, column = 1).value = i
        report = load_workbook(i)
        sheet_name = report.get_sheet_names()[1]
        report_sheet = report.get_sheet_by_name(sheet_name)
        for k in coordinate.keys():
            values = coordinate[k]
            total_sheet.cell(row = row_index, column=values[-1]).value = str(report_sheet.cell(row= values[0],column = values[1]).value)
        report.close()
        row_index += 1
    total_report.save('Total_Information.xlsx')


def report_update(report_files,name_1,report_name):
    #更新1和3的报告
    total_report = load_workbook('Total_Information.xlsx')
    total_sheet = total_report.get_sheet_by_name(name_1+'_所有报告信息')
    total_report.close()
    total_content_1 = get_report_content(total_sheet)
    n_columns = total_sheet.max_column
    total_titles = total_content_1[0:n_columns]
    special_title = []
    special_original = []
    for i in range(0,len(total_titles)):
        old_title = total_titles[i]+''
        total_titles[i],special = extract_content(total_titles[i])
        if special:
            special_title.append(total_titles[i])
            special_original.append(old_title)
    dummy_report = load_workbook('IoT-QA-D-'+report_name+'模板dummy.xlsx')
    xw_dummy_report = xw.Book('IoT-QA-D-'+report_name+'模板dummy.xlsx')
    all_sheet = dummy_report.get_sheet_names()
    for name in all_sheet[1:]:
        sheet = dummy_report.get_sheet_by_name(name)
        coordinate = get_content_coordinate(total_titles[1:],sheet)
        total_content_matrix = []
        new_titles=[]
        s_index = n_columns
        for i in range(0,int(len(total_content_1)/n_columns)-1):
            each_content = []
            each_content.extend(total_content_1[s_index:s_index+n_columns])
            new_titles.append(total_content_1[s_index])
            s_index += n_columns
            total_content_matrix.append(each_content)
        index = 0
        for i in new_titles:
            
            if i == None:
                continue
            if i not in report_files:
                create_file(i,report_name)
            xw_report = xw.Book(i)
            xw_report_sheet = xw_report.sheets(name)
            subindex = 1
            special_index = 0
            for j in total_content_matrix[index][1:]:
                if not j:
                    j = ''
                key = total_titles[subindex]+'_1'
                values = coordinate[key]
                if total_titles[subindex] in special_title:
                    
                    xw_report_sheet.range(values[0],values[1]).value = total_convert_each(special_original[special_index],str(j))
                    special_index += 1
                    subindex += 1
                    continue
                
                try:
                    test = int(str(j))
                    xw_report_sheet.range(values[0],values[1]).value = int(str(j))
                except:
                    xw_report_sheet.range(values[0],values[1]).value = str(j)
                subindex+=1
            xw_report.save()
            xw_report.close()
            index += 1

        
        
    

def get_content_coordinate(titles,sheet):
    n_rows = sheet.max_row
    n_columns = sheet.max_column    
    content_1 = get_report_content(sheet)
    coordinate = {}
    row = 2
    for i in titles:
        i += '_1'
        index_1 = content_1.index(i)
        n_row = int(index_1//n_columns+1)
        n_column = int(((index_1+1)%n_columns+1)/2)
        coordinate[i] = (n_row,n_column,row)
        row += 1
    return coordinate


def xw_get_content_coordinate(title,sheet):
    n_rows = sheet.max_row
    n_columns = sheet.max_column    
    content_1 = get_report_content(sheet)
    coordinate = {}
    row = 1
    for i in titles:
        i += '_1'
        index_1 = content_1.index(i)
        n_row = int(index_1//n_columns)
        n_column = int(((index_1)%n_columns)/2)
        coordinate[i] = (n_row,n_column,row)
        row += 1
    return coordinate
        
    
    
def extract_content(string):
    index = string.find('（')
    if index == -1:
        return string,False
    else:
        return string[0:index],True



def total_convert_each(subtitle,string):
    tick = '【√】'
    non_tick = '【】'
    if string.find('\n') != -1:
        prolix = string[string.find('\n'):]
        string = string[0:string.find('\n')]
    else:
        prolix =''
    subtitle = subtitle.split('，')
    
    if ',' in string:
        string = string.split(',')
    elif '，' in string:
        string = string.split('，')
    else:
        string = [string]
    for i in range(0,len(subtitle)):
        index = subtitle[i].index('：')
        subtitle[i] = subtitle[i][index+1:]
    new_title = ''
    for i in range(0,len(string)):
        string[i] = int(string[i])-1
    for i in range(len(subtitle)):
        if i in string:
            new_title += subtitle[i]+tick+' '
        else:
            new_title += subtitle[i]+non_tick+' '
    return new_title+prolix

def each_convert_total(subtitle):
    tick = '【√】'
    non_tick = '【】'
    subtitle = subtitle.split(' ')
    string = ''
    for i in range(0,len(subtitle)):
        if tick in subtitle[i]:
            string += (str(i+1)+',')
    return string

def create_file(name,report_name):
    source_folder=os.getcwd()
    file_list=os.listdir(source_folder)
    file = file_list[file_list.index('IoT-QA-D-'+report_name+'模板.xlsx')]
    file_path=os.path.join(source_folder,file)
    newfile_path=os.path.join(source_folder,name)
    shutil.copyfile(file_path,newfile_path)
    
    
        
def report_2_update(report_files,name_1,report_name):
    #更新里程碑测试报告
    total_report = load_workbook('Total_Information.xlsx')
    total_sheet = total_report.get_sheet_by_name(name_1+'_所有报告信息')
    total_report.close()
    total_content_1 = get_report_content(total_sheet)
    n_columns = total_sheet.max_column
    total_titles = total_content_1[0:n_columns]
    for i in range(0,len(total_titles)):
        total_titles[i],special = extract_content(total_titles[i])
    dummy_report = load_workbook('IoT-QA-D-'+report_name+'模板dummy.xlsx')
    xw_dummy_report = xw.Book('IoT-QA-D-'+report_name+'模板dummy.xlsx')
    all_sheet = dummy_report.get_sheet_names()
    dummy_report.close()
    for name in all_sheet[1:]:
        sheet = dummy_report.get_sheet_by_name(name)
        coordinate = get_content_coordinate(total_titles[1:],sheet)
        total_content_matrix = []
        new_titles=[]
        s_index = n_columns
        for i in range(0,int(len(total_content_1)/n_columns)-1):
            each_content = []
            each_content.extend(total_content_1[s_index:s_index+n_columns])
            new_titles.append(total_content_1[s_index])
            s_index += n_columns
            total_content_matrix.append(each_content)
        index = 0
        for i in new_titles:
            if i == None:
                continue
            if i not in report_files:
                create_file(i,report_name)
            xw_report = xw.Book(i)
            xw_report_sheet = xw_report.sheets(name)
            subindex = 1
            index_1 = i.find('.xlsx')
            index_2 = i.find('_')
            product_name = i[index_2:index_1]
            file_name = 'IoT-QA-D-04-V1.0小米生态链软件测试用例'+product_name+'.xlsx'
            try:
                test_data = load_workbook(file_name)
            except:
                xw_report.close()
                print(file_name)
                print('对应的该产品测试用例表格不存在！')
                continue
            for j in total_content_matrix[index][1:]:
                if not j:
                    j = ''
                key = total_titles[subindex]+'_1'
                values = coordinate[key]
                if 'Pass' in key or 'Fail' in key:
                    number = int(key[0:-2][-1])
                    if number <= 3:
                        data_sheet = test_data.get_sheet_by_name('功能性')
                        data_sheet_content = get_report_content(data_sheet)
                        data_gj = data_sheet_content.count('固件')
                        data_an = data_sheet_content.count('Android')
                        data_ios = data_sheet_content.count('iOS')
                        if number == 1:
                            s_index = 0
                            print(data_sheet_content)
                            print((data_gj+6)*data_sheet.max_column)
                            e_index = data_sheet_content.index('固件',(data_gj+5)*data_sheet.max_column)+data_sheet.max_column
                            content_gj = data_sheet_content[s_index:e_index]
                            if 'Pass' in key:
                                xw_report_sheet.range(values[0],values[1]).value = content_gj.count('PASS')
                            if 'Fail' in key:
                                xw_report_sheet.range(values[0],values[1]).value = content_gj.count('FAIL')
                        elif number == 2:
                            s_index = data_sheet_content.index('固件',(data_gj+5)*data_sheet.max_column)+data_sheet.max_column
                            e_index = data_sheet_content.index('Android',(data_gj+5+data_an)*data_sheet.max_column)+data_sheet.max_column
                            content_gj = data_sheet_content[s_index:e_index]
                            if 'Pass' in key:
                                xw_report_sheet.range(values[0],values[1]).value = content_gj.count('PASS')
                            if 'Fail' in key:
                                xw_report_sheet.range(values[0],values[1]).value = content_gj.count('FAIL')
                        else:
                            s_index = data_sheet_content.index('Android',(data_gj+5+data_an)*data_sheet.max_column)+data_sheet.max_column
                            content_gj = data_sheet_content[s_index:]
                            if 'Pass' in key:
                                xw_report_sheet.range(values[0],values[1]).value = content_gj.count('PASS')
                            if 'Fail' in key:
                                xw_report_sheet.range(values[0],values[1]).value = content_gj.count('FAIL')
                    elif number == 4:
                        data_sheet = test_data.get_sheet_by_name('可靠性')
                        data_sheet_content = get_report_content(data_sheet)
                        if 'Pass' in key:
                                xw_report_sheet.range(values[0],values[1]).value = data_sheet_content.count('PASS')
                        if 'Fail' in key:
                                xw_report_sheet.range(values[0],values[1]).value = data_sheet_content.count('FAIL')
                    elif number == 5:
                        data_sheet = test_data.get_sheet_by_name('性能')
                        data_sheet_content = get_report_content(data_sheet)
                        if 'Pass' in key:
                                xw_report_sheet.range(values[0],values[1]).value = data_sheet_content.count('PASS')
                        if 'Fail' in key:
                                xw_report_sheet.range(values[0],values[1]).value = data_sheet_content.count('FAIL')                            
                else:
                    if j:
                        try:
                            test = int(str(j))
                            xw_report_sheet.range(values[0],values[1]).value = int(str(j))
                        except:
                            xw_report_sheet.range(values[0],values[1]).value = str(j)
                    else:
                        xw_report_sheet.range(values[0],values[1]).value = ''
                subindex+=1
            test_data.close()
            xw_report.save()
            xw_report.close()
            index += 1





if __name__ == '__main__':
    switch = True
    while switch:
        which_1 = int(input('请选择操作，1:同步报告信息至总信息表，2:同步总信息表至报告，3:退出\n'))
        if which_1 == 1:
            self_update_ask()
        elif which_1 == 2:
             report_update_ask()
        elif which_1 == 3:
            break
        else:
            print('请选择一个有效输入')
